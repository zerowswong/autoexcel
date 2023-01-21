from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook('excel.xlsx')
ws = wb.active
print(wb.sheetnames)

#print B3

print(ws['B3'].value)

#change A5 to Peter

ws['A5'].value = 'Peter'
wb.save('excel.xlsx')

#Try to creat another excel file

data = [
    {
        'name': 'Alan',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name': 'Billy',
        'tall': 177,
        'age': 28,
        'weight': 90
    },
    {
        'name': 'Cathy',
        'tall': 160,
        'age': 30,
        'weight': 60
    },
    {
        'name': 'Danny',
        'tall': 155,
        'age': 50,
        'weight': 50
    },
    {
        'name': 'Elle',
        'tall': 170,
        'age': 46,
        'weight': 99
    }
]

wb = Workbook()
ws = wb.active

title = ['Name', 'Height', 'Age', 'Weight']
ws.append(title)

for person in data:
    ws.append(list(person.values()))

for col in range(2,5):
    char = get_column_letter(col)
    ws[char + '7'] = f'=AVERAGE({char + "2"}:{char + "6"})'

for col in range(1,5):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True, color="000000FF")

wb.save('data.xlsx')